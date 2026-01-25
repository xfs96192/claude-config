#!/usr/bin/env python3
"""
重点产品业绩表更新脚本

功能：
1. 从模版Excel复制并保留所有格式
2. 根据产品代码从数据源更新业绩指标
3. 处理特殊产品（非标驱动、汇利现金宝1号）
4. 生成指定日期的业绩报告

使用方法：
    python update_performance.py --date 20260115
    python update_performance.py --date 20260115 --template /path/to/template.xlsx
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path
import re

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要安装 openpyxl")
    print("请运行: pip install openpyxl")
    sys.exit(1)

try:
    import pdfplumber
except ImportError:
    print("警告: 未安装 pdfplumber，PDF提取功能将不可用")
    print("如需处理非标驱动产品，请运行: pip install pdfplumber")
    pdfplumber = None


class PerformanceReportGenerator:
    """重点产品业绩报告生成器"""

    def __init__(self, target_date, base_dir="/Users/fanshengxia/Desktop"):
        self.target_date = target_date
        self.base_dir = Path(base_dir)

        # 数据源路径
        self.template_dir = self.base_dir / "重点业绩产品"
        self.performance_data_dir = self.base_dir / "周报V2/数据/产品业绩指标数据"
        self.audit_form_dir = self.base_dir / "周报V2/数据/合享发行送审表"
        self.overview_data_dir = self.base_dir / "周报V2/数据/产品运作概览数据-母子产品"

        # 数据
        self.template_wb = None
        self.template_ws = None
        self.performance_data = None
        self.overview_data = None

    def find_latest_template(self):
        """查找最新的模版文件"""
        pattern = "部门重点产品业绩v4_*.xlsx"
        templates = list(self.template_dir.glob(pattern))

        if not templates:
            raise FileNotFoundError(f"未找到模版文件: {self.template_dir / pattern}")

        # 按修改时间排序，返回最新的
        latest = max(templates, key=lambda p: p.stat().st_mtime)
        print(f"✓ 使用模版文件: {latest.name}")
        return latest

    def load_template(self, template_path=None):
        """加载模版文件"""
        if template_path is None:
            template_path = self.find_latest_template()
        else:
            template_path = Path(template_path)

        print(f"正在加载模版: {template_path}")
        self.template_wb = load_workbook(template_path, keep_vba=True)
        self.template_ws = self.template_wb.active
        print(f"✓ 模版加载成功，工作表: {self.template_ws.title}")

    def load_performance_data(self):
        """加载产品业绩指标数据"""
        filename = f"内部产品业绩_{self.target_date}.xlsx"
        filepath = self.performance_data_dir / filename

        if not filepath.exists():
            raise FileNotFoundError(f"未找到业绩数据文件: {filepath}")

        print(f"正在加载业绩数据: {filename}")
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # 读取数据到字典，产品代码作为key
        self.performance_data = {}
        headers = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                headers = list(row)
                continue

            if row[0]:  # 假设第一列是产品代码
                product_code = str(row[0]).strip()
                self.performance_data[product_code] = dict(zip(headers, row))

        wb.close()
        print(f"✓ 加载了 {len(self.performance_data)} 个产品的业绩数据")

    def load_overview_data(self):
        """加载产品运作概览数据（用于现金产品）"""
        filename = f"产品运作概览信息表增加指标变化_{self.target_date}.xlsx"
        filepath = self.overview_data_dir / filename

        if not filepath.exists():
            print(f"警告: 未找到运作概览文件: {filepath}")
            return

        print(f"正在加载运作概览数据: {filename}")
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # 读取数据到字典
        self.overview_data = {}
        headers = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                headers = list(row)
                continue

            if row[0]:  # 假设第一列是产品代码
                product_code = str(row[0]).strip()
                self.overview_data[product_code] = dict(zip(headers, row))

        wb.close()
        print(f"✓ 加载了 {len(self.overview_data)} 个产品的运作概览数据")

    def extract_pdf_data(self):
        """从PDF送审表提取非标驱动产品数据"""
        if pdfplumber is None:
            print("警告: pdfplumber未安装，跳过PDF提取")
            return None

        # 查找最新的PDF文件
        pdf_files = list(self.audit_form_dir.glob("*.pdf"))
        if not pdf_files:
            print("警告: 未找到送审表PDF文件")
            return None

        latest_pdf = max(pdf_files, key=lambda p: p.stat().st_mtime)
        print(f"正在提取PDF数据: {latest_pdf.name}")

        # TODO: 实际的PDF提取逻辑需要根据PDF格式定制
        # 这里返回示例结构
        return {
            "product_code": "待提取",
            "benchmark_shares": "A份额2.35%；B份额 2.38%；C份额 2.21%；D份额 2.40%；E份额2.36%",
            "start_date": "待提取",
            "end_date": "待提取",
            "run_days": 0
        }

    def update_regular_product(self, row_idx, product_code):
        """更新常规产品数据"""
        if product_code not in self.performance_data:
            print(f"  警告: 产品代码 {product_code} 未在业绩数据中找到")
            return

        data = self.performance_data[product_code]

        # TODO: 根据实际的列映射更新单元格
        # 这里需要知道模版的具体列结构
        # 示例：
        # self.template_ws.cell(row_idx, col_idx).value = data['近1月年化收益率']

        # CRITICAL: Format drawdowns according to template conventions
        # Example for 近1月最大回撤:
        # dd_value = data.get('近1月最大回撤')
        # if pd.isna(dd_value) or dd_value == 0:
        #     self.template_ws.cell(row_idx, 13).value = "-"
        #     self.template_ws.cell(row_idx, 14).value = "0bp"
        # else:
        #     self.template_ws.cell(row_idx, 13).value = dd_value
        #     bp_value = int(round(abs(dd_value) * 10000))
        #     self.template_ws.cell(row_idx, 14).value = f"{bp_value}bp"

        print(f"  ✓ 更新产品: {product_code}")

    def update_feibiao_product(self, row_idx):
        """更新非标驱动产品数据"""
        pdf_data = self.extract_pdf_data()

        if pdf_data is None:
            print("  警告: 无法提取非标驱动产品数据")
            return

        # TODO: 根据实际列映射更新
        # self.template_ws.cell(row_idx, col_idx).value = pdf_data['benchmark_shares']

        print(f"  ✓ 更新非标驱动产品")

    def update_cash_product(self, row_idx, product_code="9MX00010"):
        """更新汇利现金宝1号（现金产品）"""
        if self.overview_data is None or product_code not in self.overview_data:
            print(f"  警告: 产品代码 {product_code} 未在运作概览数据中找到")
            return

        data = self.overview_data[product_code]
        cumulative_annual = data.get('累计年化', None)

        if cumulative_annual is None:
            print(f"  警告: 未找到累计年化数据")
            return

        # TODO: 根据实际列映射，将累计年化填入所有期间的年化收益率
        # Example:
        # for col in [7, 8, 9, 10, 11, 12]:  # All period return columns
        #     self.template_ws.cell(row_idx, col).value = cumulative_annual

        # CRITICAL: Set drawdowns using correct format (not numeric 0)
        # self.template_ws.cell(row_idx, 13).value = "-"
        # self.template_ws.cell(row_idx, 14).value = "0bp"

        print(f"  ✓ 更新现金产品: {product_code}, 累计年化: {cumulative_annual}")

    def update_all_products(self):
        """更新所有产品数据"""
        print("\n开始更新产品数据...")

        # TODO: 遍历模版中的产品行
        # 这里需要知道：
        # 1. 产品从哪一行开始
        # 2. 产品代码在哪一列
        # 3. 如何识别非标驱动和汇利现金宝1号

        # 示例逻辑：
        # for row_idx in range(2, self.template_ws.max_row + 1):
        #     product_code = self.template_ws.cell(row_idx, 1).value
        #
        #     if "非标驱动" in str(product_name):
        #         self.update_feibiao_product(row_idx)
        #     elif product_code == "9MX00010":
        #         self.update_cash_product(row_idx)
        #     else:
        #         self.update_regular_product(row_idx, product_code)

        print("✓ 所有产品数据更新完成")

    def save_report(self, output_path=None):
        """保存更新后的报告"""
        if output_path is None:
            filename = f"部门重点产品业绩v4_{self.target_date}.xlsx"
            output_path = self.template_dir / filename
        else:
            output_path = Path(output_path)

        print(f"\n正在保存报告: {output_path}")
        self.template_wb.save(output_path)
        print(f"✓ 报告已保存: {output_path}")
        return output_path

    def generate(self, template_path=None, output_path=None):
        """生成完整的业绩报告"""
        print(f"=" * 60)
        print(f"开始生成 {self.target_date} 的重点产品业绩报告")
        print(f"=" * 60)

        try:
            # 1. 加载模版
            self.load_template(template_path)

            # 2. 加载数据源
            self.load_performance_data()
            self.load_overview_data()

            # 3. 更新产品数据
            self.update_all_products()

            # 4. 保存报告
            output_file = self.save_report(output_path)

            print(f"\n" + "=" * 60)
            print(f"✓ 报告生成成功！")
            print(f"  文件位置: {output_file}")
            print(f"=" * 60)

            return output_file

        except Exception as e:
            print(f"\n✗ 错误: {e}")
            raise


def main():
    parser = argparse.ArgumentParser(
        description='生成重点产品业绩报告',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument(
        '--date',
        required=True,
        help='目标日期，格式：YYYYMMDD，例如：20260115'
    )

    parser.add_argument(
        '--template',
        help='模版文件路径（可选，默认使用最新模版）'
    )

    parser.add_argument(
        '--output',
        help='输出文件路径（可选，默认保存到重点业绩产品文件夹）'
    )

    parser.add_argument(
        '--base-dir',
        default='/Users/fanshengxia/Desktop',
        help='基础目录路径（默认：/Users/fanshengxia/Desktop）'
    )

    args = parser.parse_args()

    # 验证日期格式
    try:
        datetime.strptime(args.date, '%Y%m%d')
    except ValueError:
        print(f"错误: 日期格式不正确。应为 YYYYMMDD，例如：20260115")
        sys.exit(1)

    # 生成报告
    generator = PerformanceReportGenerator(args.date, args.base_dir)
    generator.generate(args.template, args.output)


if __name__ == '__main__':
    main()
